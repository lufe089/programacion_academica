"""
exports_calendario_base.py — Exportación del calendario base institucional.

Lee inputs/programacion_visual.xlsx como fuente de verdad y genera un Excel
con formato de plantilla de calendario semestral, alineado con la imagen de
referencia institucional.

Estructura de la hoja generada (una fila por semana):

    Fila 1: Título "Semestre" fusionado en todas las columnas.
    Fila 2: Encabezados — Mes | SEM | FECHA | Lunes | FECHA | <franja> | … | Domingo
    Filas 3+: Una fila por semana desde la semana de inducción hasta fin de clases.

Patrón de columnas:
    - Mes (fusionado verticalmente por mes)
    - SEM (número de semana)
    - Por cada franja: FECHA (día del mes) + columna de contenido de la franja
    - Columna Lunes (antes del primer miércoles) siempre presente
    - Columna Domingo al final (solo muestra el día, sin contenido de clase)

Fuente de datos:
    - ruta_visual_entrada: hoja 'calendario' del visual Excel ajustado manualmente
    - parametros: eventos especiales (inducción, festivos, reflexión, presencial)
    - franjas: define el orden y etiquetas de las franjas horarias
"""

import datetime
import re
from datetime import timedelta
from itertools import groupby

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import DiaSemana, Franja, Parametros


# ---------------------------------------------------------------------------
# Paleta de colores (hex sin #)
# ---------------------------------------------------------------------------

_COLOR_TITULO_BG = "1F538D"        # Azul institucional para la fila de título
_COLOR_ENCABEZADO_BG = "2E4057"    # Azul oscuro para la fila de encabezados
_COLOR_MES_BG = "2E75B6"           # Azul medio para la columna Mes
_COLOR_SEM_TEXTO = "C00000"        # Rojo oscuro para los números de semana

_COLOR_FECHA_BG = "F4F0FB"
_COLOR_LUNES_BG = "D9D9D9"
_COLOR_MIERCOLES_BG = "BDD7EE"
_COLOR_VIERNES_BG = "D9D9D9"
_COLOR_SABADO_BG = "BDD7EE"
_COLOR_DOMINGO_BG = "D9D9D9"

_COLOR_INDUCCION_BG = "F8CBAD"     #Naranja para fechas de inducción
_COLOR_FESTIVO_BG = "F8CBAD"       # Naranja para celdas de festivo
_COLOR_REFLEXION_BG = "F8CBAD"     # Naranja acero para la semana de reflexión
_COLOR_PRESENCIAL_BG = "A9D18E"    # Verde para sesiones presenciales


# ---------------------------------------------------------------------------
# Mapas de día de semana
# ---------------------------------------------------------------------------

# Nombre en español para cada DiaSemana
_NOMBRE_DIA: dict[DiaSemana, str] = {
    DiaSemana.MIERCOLES: "Miércoles",
    DiaSemana.VIERNES: "Viernes",
    DiaSemana.SABADO: "Sábado",
}

# Offset en días desde el lunes para cada DiaSemana
_OFFSET_DIA: dict[DiaSemana, int] = {
    DiaSemana.MIERCOLES: 2,   # Miércoles = lunes + 2
    DiaSemana.VIERNES: 4,     # Viernes   = lunes + 4
    DiaSemana.SABADO: 5,      # Sábado    = lunes + 5
}

# Color de fondo por defecto para columnas de contenido según el día
_COLOR_FONDO_POR_DIA: dict[DiaSemana, str] = {
    DiaSemana.MIERCOLES: _COLOR_MIERCOLES_BG,
    DiaSemana.VIERNES: _COLOR_VIERNES_BG,
    DiaSemana.SABADO: _COLOR_SABADO_BG,
}

# Nombres de los meses en español
_NOMBRE_MES: dict[int, str] = {
    1: "Enero",    2: "Febrero",   3: "Marzo",    4: "Abril",
    5: "Mayo",     6: "Junio",     7: "Julio",    8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ---------------------------------------------------------------------------
# Función pública principal
# ---------------------------------------------------------------------------

def exportar_calendario_base_desde_visual(
    ruta_visual_entrada: str,
    parametros: Parametros,
    franjas: list[Franja],
    ruta_salida: str,
) -> None:
    """
    Genera el archivo Excel de calendario base institucional.

    Lee la programación desde ruta_visual_entrada (hoja 'calendario' del
    visual Excel ajustado manualmente), enriquece con los eventos especiales
    definidos en parametros, y construye el Excel con la plantilla de
    calendario semestral.

    La primera semana del calendario es la semana que contiene FECHA_INDUCCION.
    La última semana es la semana que contiene FIN_CLASES (si está definida).

    Los números de semana en la columna SEM comienzan en parametros.semana_inicio
    (clave SEMANA_INICIO en el Excel; si no existe, por defecto es 1).

    Args:
        ruta_visual_entrada: ruta a inputs/programacion_visual.xlsx.
            Debe contener la hoja 'calendario' generada por exports_visual.py
            y potencialmente ajustada de forma manual.
        parametros: parámetros del semestre (fechas especiales, festivos, etc.).
            Se usa para eventos especiales y para delimitar el calendario.
        franjas: lista de objetos Franja en el orden definido en el Excel.
            Determina la secuencia y las etiquetas de las columnas horarias.
        ruta_salida: ruta completa del archivo Excel a generar.
    """
    contenido_visual = _leer_contenido_visual(ruta_visual_entrada, parametros)
    semanas = _calcular_semanas_del_calendario(parametros)
    semana_inicio_num = parametros.semana_inicio

    wb = Workbook()
    ws = wb.active
    ws.title = "calendario_base"

    total_columnas = _contar_columnas(franjas)

    _escribir_fila_titulo(ws, total_columnas)
    _escribir_fila_encabezados(ws, franjas)
    mes_por_fila = _escribir_filas_semanas(
        ws, semanas, semana_inicio_num, franjas, contenido_visual, parametros
    )
    _fusionar_columna_mes(ws, mes_por_fila)
    _ajustar_dimensiones(ws, franjas)

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# Lectura del visual Excel
# ---------------------------------------------------------------------------

def _leer_contenido_visual(
    ruta: str,
    parametros: Parametros,
) -> dict[datetime.date, dict[str, str]]:
    """
    Lee la hoja 'calendario' del visual Excel y extrae el contenido indexado
    por fecha del lunes de cada semana y nombre de franja.

    El visual tiene:
    - Fila 1: encabezados de semana con formato 'Sn\\ndd-Mon - dd-Mon'
    - Filas 2+: etiqueta de franja en col 1 ('NOMBRE_FRANJA\\nHH:MM ...'),
      luego contenido en las columnas de semana (o None si no hay clase)

    El contenido de cada celda tiene formato 'CODIGO (dd-Mon)' por línea.
    Esta función limpia las anotaciones de fecha y retorna solo los códigos.

    Args:
        ruta: ruta al archivo visual Excel.
        parametros: necesario para inferir el año del semestre.

    Returns:
        Diccionario {lunes_date: {nombre_franja: texto_contenido}}.
        texto_contenido contiene los códigos separados por '\\n'.
    """
    anio = parametros.inicio_clases.year
    resultado: dict[datetime.date, dict[str, str]] = {}

    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    # Paso 1: obtener la fecha del lunes para cada columna (fila 1, cols 2+)
    lunes_por_columna: dict[int, datetime.date] = {}
    for col in range(2, ws.max_column + 1):
        valor = ws.cell(1, col).value
        if valor is None:
            break
        lunes = _parsear_lunes_desde_encabezado(str(valor), anio)
        if lunes is not None:
            lunes_por_columna[col] = lunes

    # Paso 2: obtener el nombre de franja para cada fila (col 1, filas 2+)
    franja_por_fila: dict[int, str] = {}
    for fila in range(2, ws.max_row + 1):
        valor = ws.cell(fila, 1).value
        if valor is None:
            break
        nombre = _extraer_nombre_franja(str(valor))
        if nombre:
            franja_por_fila[fila] = nombre

    # Paso 3: recolectar el contenido de cada celda (franja × semana)
    for fila, nombre_franja in franja_por_fila.items():
        for col, lunes in lunes_por_columna.items():
            valor_celda = ws.cell(fila, col).value
            if valor_celda is None:
                continue
            texto = _limpiar_contenido_celda(str(valor_celda))
            if not texto:
                continue
            if lunes not in resultado:
                resultado[lunes] = {}
            resultado[lunes][nombre_franja] = texto

    wb.close()
    return resultado


def _parsear_lunes_desde_encabezado(
    texto: str,
    anio: int,
) -> datetime.date | None:
    """
    Extrae la fecha del lunes desde un encabezado de columna de semana.

    Formato esperado: 'S{n}\\n{dd-Mon} - {dd-Mon}'
    Ejemplo: 'S1\\n27-Jul - 01-Aug' → date(2026, 7, 27)

    Args:
        texto: texto del encabezado de la columna.
        anio: año del semestre para construir la fecha completa.

    Returns:
        Objeto datetime.date del lunes, o None si el formato no es reconocible.
    """
    partes = texto.strip().split("\n")
    if len(partes) < 2:
        return None

    rango = partes[1].strip()
    partes_rango = rango.split(" - ")
    if not partes_rango:
        return None

    texto_lunes = partes_rango[0].strip()
    try:
        return datetime.datetime.strptime(
            f"{texto_lunes}-{anio}", "%d-%b-%Y"
        ).date()
    except ValueError:
        return None


def _extraer_nombre_franja(texto: str) -> str | None:
    """
    Extrae el nombre de la franja desde el texto de la celda de etiqueta.

    El formato de la celda es 'NOMBRE_FRANJA\\nHH:MM … HH:MM'.
    Retorna la parte antes del primer salto de línea.

    Args:
        texto: texto de la celda (primera columna en una fila de datos).

    Returns:
        Nombre de la franja (ej. 'FRANJA_UNO_MIERCOLES'), o None si vacío
        o si es el encabezado de esquina 'Franja / Semana'.
    """
    partes = texto.strip().split("\n")
    nombre = partes[0].strip()
    if nombre and nombre != "Franja / Semana":
        return nombre
    return None


def _limpiar_contenido_celda(texto: str) -> str:
    """
    Elimina las anotaciones de fecha del contenido de una celda del visual.

    Las celdas del visual tienen formato 'CODIGO (dd-Mon)' por línea.
    Esta función extrae solo el código, eliminando la parte ' (dd-Mon)'.

    Ejemplo:
        '400CIG028 (31-Jul)\\n400FTA005 (31-Jul)' → '400CIG028\\n400FTA005'

    Args:
        texto: texto crudo de la celda del visual Excel.

    Returns:
        Texto limpio con solo los códigos, separados por '\\n'.
    """
    lineas = texto.split("\n")
    codigos = []
    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue
        # Eliminar la anotación de fecha: ' (dd-Mon)' o ' (dd-Mon, dd-Mon)'
        codigo = re.sub(r"\s*\([^)]*\)", "", linea).strip()
        if codigo:
            codigos.append(codigo)
    return "\n".join(codigos)


# ---------------------------------------------------------------------------
# Cálculo del rango de semanas del calendario
# ---------------------------------------------------------------------------

def _calcular_semanas_del_calendario(
    parametros: Parametros,
) -> list[datetime.date]:
    """
    Calcula la lista de lunes de todas las semanas del semestre.

    El calendario comienza en la semana que contiene FECHA_INDUCCION y
    termina en la semana que contiene FIN_CLASES. Si FIN_CLASES no está
    definida, se usa la semana que contiene FIN_SEMANA_SIN_CLASES como
    aproximación del final del calendario.

    Args:
        parametros: parámetros del semestre.

    Returns:
        Lista de objetos datetime.date, uno por semana (cada uno es el lunes).
    """
    lunes_inicio = _lunes_de_la_semana(parametros.fecha_induccion)

    if parametros.fin_clases is not None:
        lunes_fin = _lunes_de_la_semana(parametros.fin_clases)
    else:
        # Fallback: usar el fin de la semana sin clases como límite
        lunes_fin = _lunes_de_la_semana(parametros.fin_semana_sin_clases)

    semanas: list[datetime.date] = []
    lunes_actual = lunes_inicio
    while lunes_actual <= lunes_fin:
        semanas.append(lunes_actual)
        lunes_actual += timedelta(weeks=1)

    return semanas


def _lunes_de_la_semana(fecha: datetime.date) -> datetime.date:
    """
    Retorna el lunes de la semana que contiene la fecha dada.

    Args:
        fecha: cualquier fecha.

    Returns:
        La fecha del lunes de esa semana (puede ser igual a fecha si ya es lunes).
    """
    return fecha - timedelta(days=fecha.weekday())


# ---------------------------------------------------------------------------
# Cálculo de fechas por semana
# ---------------------------------------------------------------------------

def _fecha_de_franja_en_semana(
    lunes: datetime.date,
    franja: Franja,
) -> datetime.date:
    """
    Calcula la fecha concreta de una franja dentro de una semana dada.

    Args:
        lunes: fecha del lunes de la semana.
        franja: franja cuyo día se quiere calcular.

    Returns:
        Fecha del día de la semana correspondiente a la franja.
    """
    offset = _OFFSET_DIA[franja.dia_semana]
    return lunes + timedelta(days=offset)


def _domingo_de_la_semana(lunes: datetime.date) -> datetime.date:
    """
    Retorna el domingo de la semana que comienza el lunes dado.

    Args:
        lunes: fecha del lunes de la semana.

    Returns:
        Fecha del domingo (lunes + 6 días).
    """
    return lunes + timedelta(days=6)


# ---------------------------------------------------------------------------
# Detección de eventos especiales
# ---------------------------------------------------------------------------

# Franjas que muestran INDUCCIÓN el viernes anterior a FECHA_INDUCCION
_FRANJAS_INDUCCION_VIERNES = {"FRANJA_DOS_VIERNES", "FRANJA_TRES_VIERNES"}

# Franjas que muestran INDUCCIÓN en FECHA_INDUCCION (normalmente sábado)
_FRANJAS_INDUCCION_SABADO = {"FRANJA_UNO_SABADO", "FRANJA_DOS_SABADO"}


def _es_franja_induccion(
    franja_fecha: datetime.date,
    nombre_franja: str,
    parametros: Parametros,
) -> bool:
    """
    Indica si esta celda de franja debe mostrar INDUCCIÓN.

    Aplica en dos fechas fijas de la semana de inducción:
    - El viernes de esa semana: solo FRANJA_DOS_VIERNES y FRANJA_TRES_VIERNES.
    - FECHA_INDUCCION (normalmente sábado): solo FRANJA_UNO_SABADO y FRANJA_DOS_SABADO.

    Args:
        franja_fecha: fecha concreta del día de esta franja en la semana.
        nombre_franja: nombre identificador de la franja.
        parametros: parámetros del semestre con FECHA_INDUCCION.

    Returns:
        True si la celda debe mostrar inducción.
    """
    lunes_induccion = _lunes_de_la_semana(parametros.fecha_induccion)
    viernes_induccion = lunes_induccion + timedelta(days=4)

    if franja_fecha == viernes_induccion:
        return nombre_franja in _FRANJAS_INDUCCION_VIERNES

    if franja_fecha == parametros.fecha_induccion:
        return nombre_franja in _FRANJAS_INDUCCION_SABADO

    return False


def _es_festivo(
    fecha: datetime.date,
    parametros: Parametros,
) -> bool:
    """
    Indica si la fecha es un festivo.

    Args:
        fecha: fecha a evaluar.
        parametros: parámetros con la lista de festivos.

    Returns:
        True si la fecha está en la lista de festivos.
    """
    return fecha in parametros.festivos


def _es_semana_reflexion(
    lunes: datetime.date,
    parametros: Parametros,
) -> bool:
    """
    Indica si la semana (definida por su lunes) es la semana de reflexión.

    La comparación se hace entre el lunes de la semana evaluada y el lunes
    de la semana de INICIO_SEMANA_SIN_CLASES, para tolerar que el parámetro
    no apunte exactamente a un lunes.

    Args:
        lunes: fecha del lunes de la semana a evaluar.
        parametros: parámetros con el rango de semana sin clases.

    Returns:
        True si la semana es la semana de reflexión / semana sin clases.
    """
    lunes_reflexion = _lunes_de_la_semana(parametros.inicio_semana_sin_clases)
    lunes_fin_reflexion = _lunes_de_la_semana(parametros.fin_semana_sin_clases)
    return lunes_reflexion <= lunes <= lunes_fin_reflexion


def _es_presencial(
    fecha: datetime.date,
    parametros: Parametros,
) -> bool:
    """
    Indica si la fecha es una de las fechas designadas como presenciales.

    Args:
        fecha: fecha de la celda a evaluar.
        parametros: parámetros con las fechas presenciales opcionales.

    Returns:
        True si la fecha coincide con alguna de las 4 fechas presenciales.
    """
    fechas_presenciales = [
        parametros.viernes_presencial_uno,
        parametros.sabado_presencial_uno,
        parametros.viernes_presencial_dos,
        parametros.sabado_presencial_dos,
    ]
    return fecha in [f for f in fechas_presenciales if f is not None]


# ---------------------------------------------------------------------------
# Lógica de contenido de celdas
# ---------------------------------------------------------------------------

def _contenido_celda_franja(
    franja_fecha: datetime.date,
    nombre_franja: str,
    dia_semana: DiaSemana,
    semana_es_reflexion: bool,
    lunes: datetime.date,
    contenido_visual: dict[datetime.date, dict[str, str]],
    parametros: Parametros,
) -> tuple[str, str]:
    """
    Determina el texto y el color de fondo de una celda de franja horaria.

    Prioridad de eventos:
    1. Semana de reflexión → muestra 'SEMANA DE REFLEXIÓN' en azul acero
    2. Festivo en esa fecha → muestra 'FESTIVO' en naranja
    3. Período de inducción → muestra 'INDUCCIÓN' en dorado
    4. Presencial → muestra contenido del visual con color verde
    5. Contenido normal del visual → color por día de la semana
    6. Sin contenido → fondo suave por día de la semana

    Args:
        franja_fecha: fecha concreta del día de la semana para esta franja.
        nombre_franja: identificador de la franja (ej. 'FRANJA_UNO_VIERNES').
        dia_semana: día de la semana de la franja (para elegir el color base).
        semana_es_reflexion: True si la semana completa es de reflexión.
        lunes: fecha del lunes de la semana (índice para el visual).
        contenido_visual: contenido leído del visual Excel.
        parametros: parámetros del semestre.

    Returns:
        Tupla (texto_celda, color_fondo_hex).
    """
    color_dia = _COLOR_FONDO_POR_DIA.get(dia_semana, "FFFFFF")

    if semana_es_reflexion:
        return "SEMANA DE REFLEXIÓN", _COLOR_REFLEXION_BG

    if _es_festivo(franja_fecha, parametros):
        return "FESTIVO", _COLOR_FESTIVO_BG

    if _es_franja_induccion(franja_fecha, nombre_franja, parametros):
        return "INDUCCIÓN", _COLOR_INDUCCION_BG

    # Buscar contenido en el visual para esta semana y franja
    texto_visual = ""
    semana_visual = contenido_visual.get(lunes, {})
    if nombre_franja in semana_visual:
        texto_visual = semana_visual[nombre_franja]

    if _es_presencial(franja_fecha, parametros):
        return texto_visual, _COLOR_PRESENCIAL_BG

    return texto_visual, color_dia


def _contenido_celda_lunes(
    lunes: datetime.date,
    semana_es_reflexion: bool,
    parametros: Parametros,
) -> tuple[str, str]:
    """
    Determina el texto y el color de la celda de contenido del día Lunes.

    El lunes no tiene franjas horarias programadas, pero puede mostrar
    eventos especiales si coinciden con ese día de la semana.

    Args:
        lunes: fecha del lunes de la semana.
        semana_es_reflexion: True si la semana es de reflexión.
        parametros: parámetros del semestre.

    Returns:
        Tupla (texto_celda, color_fondo_hex).
    """
    if semana_es_reflexion:
        return "SEMANA DE REFLEXIÓN", _COLOR_REFLEXION_BG

    if _es_festivo(lunes, parametros):
        return "FESTIVO", _COLOR_FESTIVO_BG

    return "", _COLOR_LUNES_BG


# ---------------------------------------------------------------------------
# Estructura de columnas
# ---------------------------------------------------------------------------

def _contar_columnas(franjas: list[Franja]) -> int:
    """
    Calcula el número total de columnas de la hoja.

    Estructura: Mes (1) + SEM (1) + FECHA+Lunes (2) + FECHA+franja × N + Domingo (1)

    Args:
        franjas: lista de franjas del semestre.

    Returns:
        Número total de columnas.
    """
    return 2 + 2 + 2 * len(franjas) + 1


def _col_fecha_franja(franja_idx: int) -> int:
    """
    Retorna el índice de columna (base 1) de la celda FECHA antes de la franja.

    La primera franja (idx=0) tiene su FECHA en la columna 5:
    Mes(1) | SEM(2) | FECHA_Lun(3) | Lunes(4) | FECHA_franja0(5) | franja0(6) | ...

    Args:
        franja_idx: índice base 0 de la franja en la lista.

    Returns:
        Número de columna (base 1).
    """
    return 5 + 2 * franja_idx


def _col_content_franja(franja_idx: int) -> int:
    """
    Retorna el índice de columna (base 1) del contenido de la franja.

    Args:
        franja_idx: índice base 0 de la franja en la lista.

    Returns:
        Número de columna (base 1).
    """
    return 6 + 2 * franja_idx


def _col_domingo(franjas: list[Franja]) -> int:
    """
    Retorna el índice de columna (base 1) de la columna Domingo.

    Args:
        franjas: lista de franjas (para saber cuántas columnas de franjas hay).

    Returns:
        Número de columna (base 1).
    """
    return 5 + 2 * len(franjas)


def _construir_etiqueta_franja(franja: Franja) -> str:
    """
    Construye la etiqueta legible de una franja para el encabezado.

    Formato: 'Día HH:MM – HH:MM'
    Ejemplo: 'Viernes 14:00 – 17:00'

    Args:
        franja: objeto Franja con día, hora de inicio y hora de fin.

    Returns:
        Cadena de texto con el nombre del día y el rango horario.
    """
    nombre_dia = _NOMBRE_DIA.get(franja.dia_semana, franja.dia_semana.value.capitalize())
    hora_ini = franja.hora_inicio.strftime("%H:%M")
    hora_fin = franja.hora_fin.strftime("%H:%M")
    return f"{nombre_dia} {hora_ini} – {hora_fin}"


# ---------------------------------------------------------------------------
# Escritura de la hoja
# ---------------------------------------------------------------------------

def _escribir_fila_titulo(ws, total_columnas: int) -> None:
    """
    Escribe la fila de título 'Semestre' fusionada en todas las columnas.

    Args:
        ws: hoja de trabajo de openpyxl.
        total_columnas: número de columnas para determinar el rango del merge.
    """
    celda = ws.cell(row=1, column=1, value="Semestre")
    celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_TITULO_BG)
    celda.font = Font(bold=True, color="FFFFFF", size=12)
    celda.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=total_columnas,
    )


def _escribir_fila_encabezados(ws, franjas: list[Franja]) -> None:
    """
    Escribe la fila de encabezados de columnas (fila 2).

    Estructura: Mes | SEM | FECHA | Lunes | FECHA | <franja>… | Domingo

    Args:
        ws: hoja de trabajo de openpyxl.
        franjas: lista de franjas para construir los encabezados de columna.
    """
    fill_enc = PatternFill(fill_type="solid", fgColor=_COLOR_ENCABEZADO_BG)
    fuente_enc = Font(bold=True, color="FFFFFF", size=9)
    alin_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = _borde_fino()

    def _escribir_celda_enc(col: int, valor: str) -> None:
        celda = ws.cell(row=2, column=col, value=valor)
        celda.fill = fill_enc
        celda.font = fuente_enc
        celda.alignment = alin_centro
        celda.border = borde

    _escribir_celda_enc(1, "Mes")
    _escribir_celda_enc(2, "SEM")
    _escribir_celda_enc(3, "FECHA")
    _escribir_celda_enc(4, "Lunes")

    for idx, franja in enumerate(franjas):
        _escribir_celda_enc(_col_fecha_franja(idx), "FECHA")
        _escribir_celda_enc(_col_content_franja(idx), _construir_etiqueta_franja(franja))

    _escribir_celda_enc(_col_domingo(franjas), "Domingo")


def _escribir_filas_semanas(
    ws,
    semanas: list[datetime.date],
    semana_inicio_num: int,
    franjas: list[Franja],
    contenido_visual: dict[datetime.date, dict[str, str]],
    parametros: Parametros,
) -> list[tuple[str, int]]:
    """
    Escribe una fila por cada semana del calendario (filas 3 en adelante).

    Cada fila muestra: mes, número de semana, fechas por día, y el contenido
    de cada franja (asignaturas programadas o evento especial).

    Args:
        ws: hoja de trabajo de openpyxl.
        semanas: lista de lunes de cada semana del calendario.
        semana_inicio_num: número asignado a la primera semana.
        franjas: lista de franjas en orden.
        contenido_visual: {lunes: {nombre_franja: texto}} leído del visual.
        parametros: parámetros del semestre para detectar eventos especiales.

    Returns:
        Lista de tuplas (mes_nombre, numero_de_fila) para usarla en la
        fusión de la columna Mes.
    """
    borde = _borde_fino()
    alin_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    mes_por_fila: list[tuple[str, int]] = []

    for semana_idx, lunes in enumerate(semanas):
        fila = semana_idx + 3  # Las filas de datos empiezan en 3
        semana_num = semana_inicio_num + semana_idx
        nombre_mes = _NOMBRE_MES.get(lunes.month, str(lunes.month))
        semana_es_reflexion = _es_semana_reflexion(lunes, parametros)
        domingo = _domingo_de_la_semana(lunes)

        mes_por_fila.append((nombre_mes, fila))

        # Columna Mes (el valor se escribe ahora; la fusión se aplica después)
        celda_mes = ws.cell(row=fila, column=1, value=nombre_mes)
        celda_mes.font = Font(bold=True, color="FFFFFF", size=10)
        celda_mes.alignment = alin_centro
        celda_mes.border = borde

        # Columna SEM (número de semana en rojo)
        celda_sem = ws.cell(row=fila, column=2, value=semana_num)
        celda_sem.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        celda_sem.font = Font(bold=True, color=_COLOR_SEM_TEXTO, size=9)
        celda_sem.alignment = alin_centro
        celda_sem.border = borde

        # FECHA del lunes
        celda_fecha_lunes = ws.cell(row=fila, column=3, value=lunes.day)
        celda_fecha_lunes.fill = PatternFill(fill_type="solid", fgColor=_COLOR_FECHA_BG)
        celda_fecha_lunes.font = Font(size=9)
        celda_fecha_lunes.alignment = alin_centro
        celda_fecha_lunes.border = borde

        # Contenido del lunes (sin franjas programadas, solo eventos especiales)
        texto_lunes, color_lunes = _contenido_celda_lunes(lunes, semana_es_reflexion, parametros)
        celda_lunes = ws.cell(row=fila, column=4, value=texto_lunes)
        celda_lunes.fill = PatternFill(fill_type="solid", fgColor=color_lunes)
        celda_lunes.font = Font(size=8)
        celda_lunes.alignment = alin_centro
        celda_lunes.border = borde

        # FECHA + contenido de cada franja
        for idx, franja in enumerate(franjas):
            franja_fecha = _fecha_de_franja_en_semana(lunes, franja)
            col_fecha = _col_fecha_franja(idx)
            col_content = _col_content_franja(idx)

            # Celda FECHA de la franja
            celda_fecha_f = ws.cell(row=fila, column=col_fecha, value=franja_fecha.day)
            celda_fecha_f.fill = PatternFill(fill_type="solid", fgColor=_COLOR_FECHA_BG)
            celda_fecha_f.font = Font(size=9)
            celda_fecha_f.alignment = alin_centro
            celda_fecha_f.border = borde

            # Celda de contenido de la franja
            texto, color = _contenido_celda_franja(
                franja_fecha=franja_fecha,
                nombre_franja=franja.nombre,
                dia_semana=franja.dia_semana,
                semana_es_reflexion=semana_es_reflexion,
                lunes=lunes,
                contenido_visual=contenido_visual,
                parametros=parametros,
            )
            celda_cont = ws.cell(row=fila, column=col_content, value=texto)
            celda_cont.fill = PatternFill(fill_type="solid", fgColor=color)
            celda_cont.font = Font(size=8)
            celda_cont.alignment = alin_centro
            celda_cont.border = borde

        # Columna Domingo (solo fecha, sin contenido)
        col_dom = _col_domingo(franjas)
        celda_dom = ws.cell(row=fila, column=col_dom, value=domingo.day)
        celda_dom.fill = PatternFill(fill_type="solid", fgColor=_COLOR_DOMINGO_BG)
        celda_dom.font = Font(size=9)
        celda_dom.alignment = alin_centro
        celda_dom.border = borde

    return mes_por_fila


# ---------------------------------------------------------------------------
# Fusión de la columna Mes
# ---------------------------------------------------------------------------

def _fusionar_columna_mes(
    ws,
    mes_por_fila: list[tuple[str, int]],
) -> None:
    """
    Fusiona verticalmente las celdas de la columna Mes para semanas del mismo mes.

    Para cada grupo de semanas consecutivas del mismo mes, une las celdas de la
    columna 1 en un solo bloque y aplica el estilo de mes.

    Args:
        ws: hoja de trabajo de openpyxl.
        mes_por_fila: lista de (mes_nombre, fila) en orden secuencial, tal como
            fue retornada por _escribir_filas_semanas.
    """
    fill_mes = PatternFill(fill_type="solid", fgColor=_COLOR_MES_BG)
    fuente_mes = Font(bold=True, color="FFFFFF", size=10)
    alin_mes = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = _borde_fino()

    # Agrupar filas consecutivas del mismo mes
    for mes, grupo_iter in groupby(mes_por_fila, key=lambda par: par[0]):
        grupo = list(grupo_iter)
        primera_fila = grupo[0][1]
        ultima_fila = grupo[-1][1]

        # Fusionar si hay más de una fila en el grupo
        if primera_fila < ultima_fila:
            ws.merge_cells(
                start_row=primera_fila, start_column=1,
                end_row=ultima_fila, end_column=1,
            )

        # Aplicar estilo a la celda superior (única visible tras la fusión)
        celda = ws.cell(row=primera_fila, column=1, value=mes)
        celda.fill = fill_mes
        celda.font = fuente_mes
        celda.alignment = alin_mes
        celda.border = borde


# ---------------------------------------------------------------------------
# Ajuste de dimensiones y navegación
# ---------------------------------------------------------------------------

def _ajustar_dimensiones(ws, franjas: list[Franja]) -> None:
    """
    Ajusta los anchos de columna, altos de fila y congela los paneles.

    Args:
        ws: hoja de trabajo de openpyxl.
        franjas: lista de franjas para saber cuántas columnas de datos hay.
    """
    # Anchos de columnas fijas
    ws.column_dimensions["A"].width = 11   # Mes
    ws.column_dimensions["B"].width = 5    # SEM
    ws.column_dimensions["C"].width = 7    # FECHA Lunes
    ws.column_dimensions["D"].width = 14   # Lunes contenido

    # Anchos para las columnas FECHA + franja
    for idx in range(len(franjas)):
        col_fecha = get_column_letter(_col_fecha_franja(idx))
        col_cont = get_column_letter(_col_content_franja(idx))
        ws.column_dimensions[col_fecha].width = 7
        ws.column_dimensions[col_cont].width = 16

    # Domingo
    col_dom = get_column_letter(_col_domingo(franjas))
    ws.column_dimensions[col_dom].width = 8

    # Altos de fila
    ws.row_dimensions[1].height = 22   # Título
    ws.row_dimensions[2].height = 32   # Encabezados (texto puede tener wrap)
    for fila in range(3, ws.max_row + 1):
        ws.row_dimensions[fila].height = 22

    # Congelar paneles: la navegación horizontal/vertical deja fijos Mes y SEM
    ws.freeze_panes = "C3"


# ---------------------------------------------------------------------------
# Utilidades de estilo
# ---------------------------------------------------------------------------

def _borde_fino() -> Border:
    """
    Retorna un objeto Border con líneas finas en los cuatro lados.

    Returns:
        Border de openpyxl con estilo 'thin' en gris suave.
    """
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)
