"""
exports_matriz.py — Exportación de la matriz de horas de la programación.

Genera un archivo Excel con una matriz donde:

    - Filas: asignaturas programadas.
    - Columnas: fechas del calendario (miércoles, viernes, sábados).
    - Celdas: número de horas programadas para esa asignatura en esa fecha.

Incluye:
    - Filas auxiliares de contexto (semana, día de la semana).
    - Fila resumen de horas efectivas por fecha.
    - Columnas finales de totales y estado por asignatura.
    - Formato visual para fechas especiales (presenciales, inducción, días sin clase).

Esta salida está diseñada para revisión humana de la distribución de carga.
"""

import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exports_hours import calcular_numero_semana
from models import Asignatura, Parametros, RestriccionProgramacion


# ---------------------------------------------------------------------------
# Colores de formato visual
# ---------------------------------------------------------------------------

_COLOR_ENCABEZADO = "2F4F8F"           # azul oscuro
_COLOR_FILA_AUXILIAR = "E8E8E8"        # gris claro
_COLOR_FECHA_INDUCCION = "FFD700"      # dorado
_COLOR_FECHA_PRESENCIAL = "90EE90"     # verde claro (para semanas presenciales)
_COLOR_FECHA_SIN_CLASE = "FFCCCB"      # rojo claro
_COLOR_SEMANA_PAR = "E8E8E8"           # gris claro para semanas pares
_COLOR_SEMANA_IMPAR = "FFFFFF"         # blanco para semanas impares
_COLOR_RESUMEN_OK = "C6EFCE"           # verde pastel
_COLOR_RESUMEN_INCOMPLETA = "FFC7CE"   # rojo pastel
_COLOR_RESUMEN_EXCEDE = "FFEB9C"       # amarillo pastel


# ---------------------------------------------------------------------------
# Constantes para identificar grupos de franja compartida
# ---------------------------------------------------------------------------

# Restricciones que implican que todas las asignaturas del tipo comparten franja
_RESTRICCIONES_FRANJA_COMPARTIDA = {
    RestriccionProgramacion.MISMA_FRANJA,
    RestriccionProgramacion.OBLIGATORIOS_MISMA_FRANJA,
}


# ---------------------------------------------------------------------------
# Función principal de exportación
# ---------------------------------------------------------------------------

def exportar_matriz_horas(
    sesiones: pd.DataFrame,
    asignaturas: list[Asignatura],
    parametros: Parametros,
    calendario: pd.DataFrame,
    ruta_salida: str,
) -> None:
    """
    Exporta la programación en formato de matriz asignaturas × fechas.

    Genera un archivo Excel con la hoja 'matriz' que permite revisar
    visualmente la distribución de horas por asignatura y fecha.

    Args:
        sesiones: DataFrame de sesiones asignadas (salida de asignar_sesiones).
        asignaturas: lista de asignaturas de la corrida actual.
        parametros: parámetros de la corrida (incluye fechas especiales).
        calendario: DataFrame del calendario con todas las fechas del periodo.
        ruta_salida: ruta completa del archivo Excel a generar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "matriz"

    if sesiones.empty:
        wb.save(ruta_salida)
        return

    # Preparar datos
    df = sesiones.copy()
    df["semana"] = df["fecha"].apply(
        lambda fecha: calcular_numero_semana(fecha, parametros.inicio_clases)
    )

    # Obtener TODAS las fechas del calendario (incluyendo las no programables)
    todas_las_fechas = sorted(calendario["fecha"].unique())
    fechas_especiales = _construir_fechas_especiales(parametros, calendario)

    # Construir mapas auxiliares para el sombreado por semana
    semana_por_fecha = _construir_semana_por_fecha(todas_las_fechas, parametros)
    semanas_presenciales = _identificar_semanas_presenciales(parametros)

    # Construir matriz de horas por (asignatura, fecha)
    matriz_horas = _construir_matriz_horas(df, asignaturas, todas_las_fechas)

    # Construir grupos para la fórmula de horas efectivas
    grupos_max, filas_suma = _construir_grupos_formula_resumen(asignaturas)

    # Escribir contenido
    _escribir_filas_auxiliares(ws, todas_las_fechas, parametros, fechas_especiales, semana_por_fecha, semanas_presenciales)
    _escribir_encabezado_asignaturas(ws, todas_las_fechas, semana_por_fecha, semanas_presenciales, fechas_especiales)
    _escribir_datos_asignaturas(ws, asignaturas, matriz_horas, todas_las_fechas, semana_por_fecha, semanas_presenciales, fechas_especiales)
    _escribir_fila_resumen_carga(ws, df, todas_las_fechas, len(asignaturas), semana_por_fecha, semanas_presenciales, fechas_especiales, grupos_max, filas_suma)
    _escribir_columnas_totales(ws, asignaturas, matriz_horas, todas_las_fechas)

    # Formato final
    _ajustar_dimensiones(ws, todas_las_fechas, asignaturas)
    _congelar_paneles(ws)

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# Construcción de datos auxiliares
# ---------------------------------------------------------------------------

def _construir_fechas_especiales(
    parametros: Parametros,
    calendario: pd.DataFrame,
) -> dict[datetime.date, str]:
    """
    Construye un diccionario de fechas especiales y su tipo.

    Tipos posibles: 'induccion', 'sin_clase', 'presencial'.

    Las fechas no programables del calendario (festivos, semana sin clases)
    se marcan como 'sin_clase'.

    Args:
        parametros: parámetros de la corrida con fechas especiales.
        calendario: DataFrame del calendario con columna 'es_programable'.

    Returns:
        Diccionario fecha → tipo de fecha especial.
    """
    fechas = {}

    # Marcar todas las fechas no programables del calendario como 'sin_clase'
    fechas_no_programables = calendario[calendario["es_programable"] == False]["fecha"].unique()
    for fecha in fechas_no_programables:
        fechas[fecha] = "sin_clase"

    # Fecha de inducción (prioridad más alta, sobrescribe sin_clase)
    if parametros.fecha_induccion:
        fechas[parametros.fecha_induccion] = "induccion"

    # Fechas presenciales (prioridad menor que inducción y sin_clase)
    # Solo aplica si la fecha NO está ya marcada como sin_clase o inducción
    fechas_presenciales = [
        parametros.viernes_presencial_uno,
        parametros.sabado_presencial_uno,
        parametros.viernes_presencial_dos,
        parametros.sabado_presencial_dos,
    ]
    for fecha in fechas_presenciales:
        if fecha and fecha not in fechas:
            fechas[fecha] = "presencial"

    return fechas


def _construir_matriz_horas(
    df: pd.DataFrame,
    asignaturas: list[Asignatura],
    fechas: list[datetime.date],
) -> dict[tuple[str, datetime.date], float]:
    """
    Construye un diccionario (codigo, fecha) → horas programadas.

    Args:
        df: DataFrame de sesiones con columnas 'codigo', 'fecha', 'horas_sesion'.
        asignaturas: lista de asignaturas.
        fechas: lista de fechas ordenadas.

    Returns:
        Diccionario con las horas por cada combinación (asignatura, fecha).
    """
    matriz = {}

    # Inicializar todas las celdas en 0
    for asignatura in asignaturas:
        for fecha in fechas:
            matriz[(asignatura.codigo, fecha)] = 0.0

    # Llenar con las horas reales
    for _, fila in df.iterrows():
        clave = (fila["codigo"], fila["fecha"])
        if clave in matriz:
            matriz[clave] += fila["horas_sesion"]

    return matriz


def _construir_semana_por_fecha(
    fechas: list[datetime.date],
    parametros: Parametros,
) -> dict[datetime.date, int]:
    """
    Construye un diccionario fecha → número de semana.

    Args:
        fechas: lista de fechas ordenadas.
        parametros: parámetros para calcular número de semana.

    Returns:
        Diccionario fecha → número de semana.
    """
    return {
        fecha: calcular_numero_semana(fecha, parametros.inicio_clases)
        for fecha in fechas
    }


def _identificar_semanas_presenciales(parametros: Parametros) -> set[int]:
    """
    Identifica los números de semana que tienen encuentro presencial.

    Una semana se considera presencial si contiene al menos una fecha
    de viernes o sábado presencial.

    Args:
        parametros: parámetros de la corrida con fechas presenciales.

    Returns:
        Conjunto de números de semana con encuentro presencial.
    """
    semanas = set()

    fechas_presenciales = [
        parametros.viernes_presencial_uno,
        parametros.sabado_presencial_uno,
        parametros.viernes_presencial_dos,
        parametros.sabado_presencial_dos,
    ]

    for fecha in fechas_presenciales:
        if fecha:
            semana = calcular_numero_semana(fecha, parametros.inicio_clases)
            semanas.add(semana)

    return semanas


def _obtener_color_semana(
    semana: int,
    semanas_presenciales: set[int],
) -> str:
    """
    Retorna el color de fondo según el número de semana.

    - Semanas con encuentro presencial: verde claro
    - Semanas pares: gris claro
    - Semanas impares: blanco

    Args:
        semana: número de semana.
        semanas_presenciales: conjunto de semanas con encuentro presencial.

    Returns:
        Código de color hexadecimal.
    """
    if semana in semanas_presenciales:
        return _COLOR_FECHA_PRESENCIAL
    elif semana % 2 == 0:
        return _COLOR_SEMANA_PAR
    else:
        return _COLOR_SEMANA_IMPAR


def _construir_grupos_formula_resumen(
    asignaturas: list[Asignatura],
) -> tuple[dict[str, list[int]], list[int]]:
    """
    Construye los grupos de filas para la fórmula de horas efectivas.

    Las asignaturas con restricción de franja compartida (Obligatorio,
    TemasAvanzados) se agrupan por tipo: para cada grupo se usará MAX()
    en la fórmula, ya que comparten franja y no deben sumarse.

    Las asignaturas individuales (ProcesoDesarrollo, etc.) se suman
    directamente ya que cada una ocupa su propia franja.

    Args:
        asignaturas: lista de asignaturas en el orden en que se escriben.

    Returns:
        Tupla (grupos_max, filas_suma) donde:
        - grupos_max: dict tipo → lista de filas Excel del grupo
        - filas_suma: lista de filas Excel que se suman individualmente
    """
    fila_inicio = 5  # Los datos de asignaturas empiezan en fila 5

    grupos_max: dict[str, list[int]] = {}
    filas_suma: list[int] = []

    for idx, asignatura in enumerate(asignaturas):
        fila_excel = fila_inicio + idx

        if asignatura.restriccion_programacion in _RESTRICCIONES_FRANJA_COMPARTIDA:
            # Asignaturas que comparten franja: agrupar por tipo
            tipo = asignatura.tipo
            if tipo not in grupos_max:
                grupos_max[tipo] = []
            grupos_max[tipo].append(fila_excel)
        else:
            # Asignaturas individuales: se suman directamente
            filas_suma.append(fila_excel)

    return grupos_max, filas_suma


# ---------------------------------------------------------------------------
# Escritura de filas auxiliares (semana, día)
# ---------------------------------------------------------------------------

def _escribir_filas_auxiliares(
    ws,
    fechas: list[datetime.date],
    parametros: Parametros,
    fechas_especiales: dict[datetime.date, str],
    semana_por_fecha: dict[datetime.date, int],
    semanas_presenciales: set[int],
) -> None:
    """
    Escribe las filas auxiliares de contexto: semana y día de la semana.

    Fila 1: número de semana
    Fila 2: día de la semana (Mié, Vie, Sáb)
    Fila 3: fecha (dd/mm)

    Aplica sombreado alternado por semana (blanco/gris) y verde para semanas presenciales.

    Args:
        ws: hoja de trabajo de openpyxl.
        fechas: lista de fechas ordenadas.
        parametros: parámetros para calcular número de semana.
        fechas_especiales: diccionario de fechas con formato especial.
        semana_por_fecha: diccionario fecha → número de semana.
        semanas_presenciales: conjunto de semanas con encuentro presencial.
    """
    estilo_auxiliar = PatternFill(fill_type="solid", fgColor=_COLOR_FILA_AUXILIAR)
    fuente_auxiliar = Font(size=8, bold=True)
    alineacion = Alignment(horizontal="center", vertical="center")
    borde = _crear_borde_fino()

    dias_semana = {
        2: "Mié",
        4: "Vie",
        5: "Sáb",
    }

    # Columnas A y B: etiquetas (Tipo y Asignatura)
    ws.cell(row=1, column=1, value="").font = fuente_auxiliar
    ws.cell(row=2, column=1, value="").font = fuente_auxiliar
    ws.cell(row=3, column=1, value="").font = fuente_auxiliar
    ws.cell(row=1, column=2, value="Semana").font = fuente_auxiliar
    ws.cell(row=2, column=2, value="Día").font = fuente_auxiliar
    ws.cell(row=3, column=2, value="Fecha").font = fuente_auxiliar

    for i in range(1, 4):
        celda_a = ws.cell(row=i, column=1)
        celda_a.fill = estilo_auxiliar
        celda_a.alignment = alineacion
        celda_a.border = borde

        celda_b = ws.cell(row=i, column=2)
        celda_b.fill = estilo_auxiliar
        celda_b.alignment = alineacion
        celda_b.border = borde

    # Columnas de fechas (empiezan en columna C)
    for col_idx, fecha in enumerate(fechas):
        columna = col_idx + 3  # Columna C en adelante

        semana = semana_por_fecha[fecha]
        dia = dias_semana.get(fecha.weekday(), "")
        fecha_texto = fecha.strftime("%d/%m")

        # Determinar color de fondo: fechas especiales tienen prioridad sobre sombreado de semana
        tipo_fecha = fechas_especiales.get(fecha)

        if tipo_fecha == "induccion":
            color_fondo = _COLOR_FECHA_INDUCCION
        elif tipo_fecha == "sin_clase":
            color_fondo = _COLOR_FECHA_SIN_CLASE
        else:
            # Usar sombreado por semana (presencial/par/impar)
            color_fondo = _obtener_color_semana(semana, semanas_presenciales)

        for fila, valor in [(1, f"S{semana}"), (2, dia), (3, fecha_texto)]:
            celda = ws.cell(row=fila, column=columna, value=valor)
            celda.font = fuente_auxiliar
            celda.alignment = alineacion
            celda.border = borde
            celda.fill = PatternFill(fill_type="solid", fgColor=color_fondo)


# ---------------------------------------------------------------------------
# Escritura del encabezado de asignaturas
# ---------------------------------------------------------------------------

def _escribir_encabezado_asignaturas(
    ws,
    fechas: list[datetime.date],
    semana_por_fecha: dict[datetime.date, int],
    semanas_presenciales: set[int],
    fechas_especiales: dict[datetime.date, str],
) -> None:
    """
    Escribe el encabezado de las columnas Tipo y Asignatura en la fila 4.

    También aplica sombreado por semana a las celdas de la fila 4 sobre las fechas.
    Las fechas sin clase se marcan en rojo.

    Args:
        ws: hoja de trabajo de openpyxl.
        fechas: lista de fechas ordenadas.
        semana_por_fecha: diccionario fecha → número de semana.
        semanas_presenciales: conjunto de semanas con encuentro presencial.
        fechas_especiales: diccionario fecha → tipo de fecha especial.
    """
    estilo_encabezado = PatternFill(fill_type="solid", fgColor=_COLOR_ENCABEZADO)
    fuente_encabezado = Font(bold=True, color="FFFFFF", size=9)
    alineacion = Alignment(horizontal="left", vertical="center")
    borde = _crear_borde_fino()

    # Columna A: Tipo
    celda_tipo = ws.cell(row=4, column=1, value="Tipo")
    celda_tipo.fill = estilo_encabezado
    celda_tipo.font = fuente_encabezado
    celda_tipo.alignment = alineacion
    celda_tipo.border = borde

    # Columna B: Asignatura
    celda_asig = ws.cell(row=4, column=2, value="Asignatura")
    celda_asig.fill = estilo_encabezado
    celda_asig.font = fuente_encabezado
    celda_asig.alignment = alineacion
    celda_asig.border = borde

    # Aplicar sombreado por semana a las columnas de fechas en fila 4 (empiezan en C)
    for col_idx, fecha in enumerate(fechas):
        columna = col_idx + 3
        semana = semana_por_fecha[fecha]

        # Determinar color: sin_clase tiene prioridad (rojo)
        tipo_fecha = fechas_especiales.get(fecha)
        if tipo_fecha == "sin_clase":
            color_fondo = _COLOR_FECHA_SIN_CLASE
        elif tipo_fecha == "induccion":
            color_fondo = _COLOR_FECHA_INDUCCION
        else:
            color_fondo = _obtener_color_semana(semana, semanas_presenciales)

        celda = ws.cell(row=4, column=columna)
        celda.fill = PatternFill(fill_type="solid", fgColor=color_fondo)
        celda.border = borde


# ---------------------------------------------------------------------------
# Escritura de datos de asignaturas
# ---------------------------------------------------------------------------

def _escribir_datos_asignaturas(
    ws,
    asignaturas: list[Asignatura],
    matriz_horas: dict[tuple[str, datetime.date], float],
    fechas: list[datetime.date],
    semana_por_fecha: dict[datetime.date, int],
    semanas_presenciales: set[int],
    fechas_especiales: dict[datetime.date, str],
) -> None:
    """
    Escribe las filas de datos: una por asignatura con las horas por fecha.

    Aplica sombreado alternado por semana (blanco/gris) y verde para semanas presenciales.

    Args:
        ws: hoja de trabajo de openpyxl.
        asignaturas: lista de asignaturas.
        matriz_horas: diccionario (codigo, fecha) → horas.
        fechas: lista de fechas ordenadas.
        semana_por_fecha: diccionario fecha → número de semana.
        semanas_presenciales: conjunto de semanas con encuentro presencial.
        fechas_especiales: diccionario de fechas con formato especial.
    """
    fuente_asignatura = Font(size=8, bold=True)
    fuente_horas = Font(size=8)
    alineacion_izq = Alignment(horizontal="left", vertical="center")
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    borde = _crear_borde_fino()

    fila_inicio = 5  # Datos empiezan en fila 5

    for idx, asignatura in enumerate(asignaturas):
        fila = fila_inicio + idx

        # Columna A: Tipo de asignatura
        celda_tipo = ws.cell(row=fila, column=1, value=asignatura.tipo)
        celda_tipo.font = fuente_asignatura
        celda_tipo.alignment = alineacion_izq
        celda_tipo.border = borde

        # Columna B: Nombre de asignatura (solo el nombre, sin código)
        celda_nombre = ws.cell(row=fila, column=2, value=asignatura.nombre)
        celda_nombre.font = fuente_asignatura
        celda_nombre.alignment = alineacion_izq
        celda_nombre.border = borde

        # Columnas de fechas: horas programadas (empiezan en columna C)
        for col_idx, fecha in enumerate(fechas):
            columna = col_idx + 3
            horas = matriz_horas.get((asignatura.codigo, fecha), 0.0)
            semana = semana_por_fecha[fecha]

            celda = ws.cell(row=fila, column=columna)
            celda.border = borde
            celda.alignment = alineacion_centro
            celda.font = fuente_horas

            # Determinar color de fondo
            tipo_fecha = fechas_especiales.get(fecha)
            if tipo_fecha == "induccion":
                color_fondo = _COLOR_FECHA_INDUCCION
            elif tipo_fecha == "sin_clase":
                color_fondo = _COLOR_FECHA_SIN_CLASE
            else:
                color_fondo = _obtener_color_semana(semana, semanas_presenciales)

            celda.fill = PatternFill(fill_type="solid", fgColor=color_fondo)

            if horas > 0:
                # Mostrar como entero si no tiene decimales
                if horas == int(horas):
                    celda.value = int(horas)
                else:
                    celda.value = round(horas, 1)


# ---------------------------------------------------------------------------
# Escritura de fila resumen de carga diaria
# ---------------------------------------------------------------------------

def _escribir_fila_resumen_carga(
    ws,
    df: pd.DataFrame,
    fechas: list[datetime.date],
    num_asignaturas: int,
    semana_por_fecha: dict[datetime.date, int],
    semanas_presenciales: set[int],
    fechas_especiales: dict[datetime.date, str],
    grupos_max: dict[str, list[int]],
    filas_suma: list[int],
) -> None:
    """
    Escribe la fila de resumen con horas efectivas por fecha usando fórmulas Excel.

    Las horas efectivas se calculan mediante fórmulas que respetan la regla
    de franja compartida:
    - Grupos con franja compartida (Obligatorio, TemasAvanzados): se usa MAX()
      para tomar solo una vez las horas de la franja.
    - Asignaturas individuales (ProcesoDesarrollo, etc.): se suman directamente.

    La fórmula resultante por columna es del tipo:
    =MAX(C5:C7)+MAX(C11:C13)+C8+C9+C10

    Args:
        ws: hoja de trabajo de openpyxl.
        df: DataFrame de sesiones (usado para calcular el color de la celda).
        fechas: lista de fechas ordenadas.
        num_asignaturas: número de asignaturas para calcular la fila.
        semana_por_fecha: diccionario fecha → número de semana.
        semanas_presenciales: conjunto de semanas con encuentro presencial.
        fechas_especiales: diccionario fecha → tipo de fecha especial.
        grupos_max: diccionario tipo → lista de filas Excel del grupo.
        filas_suma: lista de filas Excel que se suman individualmente.
    """
    fila_resumen = 5 + num_asignaturas + 1  # Una fila vacía de separación

    estilo_resumen = PatternFill(fill_type="solid", fgColor=_COLOR_ENCABEZADO)
    fuente_resumen = Font(bold=True, color="FFFFFF", size=8)
    alineacion = Alignment(horizontal="center", vertical="center")
    borde = _crear_borde_fino()

    # Etiqueta en columnas A y B
    celda_etiqueta_a = ws.cell(row=fila_resumen, column=1, value="")
    celda_etiqueta_a.fill = estilo_resumen
    celda_etiqueta_a.font = fuente_resumen
    celda_etiqueta_a.alignment = alineacion
    celda_etiqueta_a.border = borde

    celda_etiqueta_b = ws.cell(row=fila_resumen, column=2, value="Horas efectivas")
    celda_etiqueta_b.fill = estilo_resumen
    celda_etiqueta_b.font = fuente_resumen
    celda_etiqueta_b.alignment = alineacion
    celda_etiqueta_b.border = borde

    # Generar fórmula y escribir celdas por fecha
    for col_idx, fecha in enumerate(fechas):
        columna = col_idx + 3  # Columna C en adelante
        semana = semana_por_fecha[fecha]
        letra_columna = get_column_letter(columna)

        # Construir la fórmula para esta columna
        formula = _construir_formula_horas_efectivas(letra_columna, grupos_max, filas_suma)

        # Calcular horas efectivas en Python solo para determinar el color
        sesiones_fecha = df[df["fecha"] == fecha]
        if sesiones_fecha.empty:
            horas_efectivas = 0
        else:
            horas_por_franja = sesiones_fecha.groupby("nombre_franja")["horas_sesion"].first()
            horas_efectivas = horas_por_franja.sum()

        celda = ws.cell(row=fila_resumen, column=columna)
        celda.border = borde
        celda.alignment = alineacion
        celda.font = Font(size=8, bold=True)

        # Verificar si es fecha especial (sin clase o inducción)
        tipo_fecha = fechas_especiales.get(fecha)

        if tipo_fecha == "sin_clase":
            # Día sin clase: color rojo, sin fórmula
            celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_FECHA_SIN_CLASE)
        elif tipo_fecha == "induccion":
            # Día de inducción: color dorado, sin fórmula
            celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_FECHA_INDUCCION)
        else:
            # Escribir la fórmula
            celda.value = formula

            # Color según carga (viernes ~7h, sábado ~10h)
            dia_semana = fecha.weekday()
            if horas_efectivas > 0:
                if dia_semana == 4:  # Viernes
                    if horas_efectivas <= 8:
                        celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_RESUMEN_OK)
                    else:
                        celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_RESUMEN_EXCEDE)
                elif dia_semana == 5:  # Sábado
                    if horas_efectivas <= 10:
                        celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_RESUMEN_OK)
                    else:
                        celda.fill = PatternFill(fill_type="solid", fgColor=_COLOR_RESUMEN_EXCEDE)
                else:
                    # Miércoles u otro día: usar color base de semana
                    color_base = _obtener_color_semana(semana, semanas_presenciales)
                    celda.fill = PatternFill(fill_type="solid", fgColor=color_base)
            else:
                # Sin horas: usar color base de semana
                color_base = _obtener_color_semana(semana, semanas_presenciales)
                celda.fill = PatternFill(fill_type="solid", fgColor=color_base)


def _construir_formula_horas_efectivas(
    letra_columna: str,
    grupos_max: dict[str, list[int]],
    filas_suma: list[int],
) -> str:
    """
    Construye la fórmula Excel para calcular horas efectivas de una columna.

    La fórmula combina:
    - MAX() para cada grupo de franja compartida (ej: MAX(C5:C7) para Obligatorio)
    - Referencias directas para asignaturas individuales (ej: +C8+C9)

    Args:
        letra_columna: letra de la columna Excel (ej: 'C', 'D', 'AA').
        grupos_max: diccionario tipo → lista de filas Excel del grupo.
        filas_suma: lista de filas Excel que se suman individualmente.

    Returns:
        Fórmula Excel como string (ej: '=MAX(C5:C7)+MAX(C11:C13)+C8+C9').
    """
    partes = []

    # Agregar MAX() por cada grupo de franja compartida
    for tipo, filas in grupos_max.items():
        if len(filas) == 1:
            # Un solo miembro: referencia directa
            partes.append(f"{letra_columna}{filas[0]}")
        else:
            # Múltiples miembros: usar MAX del rango
            fila_min = min(filas)
            fila_max = max(filas)
            partes.append(f"MAX({letra_columna}{fila_min}:{letra_columna}{fila_max})")

    # Agregar referencias directas para asignaturas individuales
    for fila in filas_suma:
        partes.append(f"{letra_columna}{fila}")

    if not partes:
        return "=0"

    return "=" + "+".join(partes)


# ---------------------------------------------------------------------------
# Escritura de columnas de totales por asignatura
# ---------------------------------------------------------------------------

def _escribir_columnas_totales(
    ws,
    asignaturas: list[Asignatura],
    matriz_horas: dict[tuple[str, datetime.date], float],
    fechas: list[datetime.date],
) -> None:
    """
    Escribe las columnas finales con totales y estado por asignatura.

    Columnas: Asignadas (fórmula SUMA) | Objetivo | Diferencia (fórmula) | Estado (fórmula SI)

    Las columnas Asignadas, Diferencia y Estado usan fórmulas de Excel
    para permitir verificación y ajustes manuales por el usuario.

    Args:
        ws: hoja de trabajo de openpyxl.
        asignaturas: lista de asignaturas.
        matriz_horas: diccionario (codigo, fecha) → horas.
        fechas: lista de fechas ordenadas.
    """
    col_inicio = len(fechas) + 3  # Después de la última fecha (columnas A, B + fechas)
    col_asignadas = col_inicio
    col_objetivo = col_inicio + 1
    col_diferencia = col_inicio + 2
    col_estado = col_inicio + 3

    # Letra de la primera y última columna de fechas (para fórmulas SUMA)
    col_primera_fecha = get_column_letter(3)  # Columna C
    col_ultima_fecha = get_column_letter(len(fechas) + 2)

    estilo_encabezado = PatternFill(fill_type="solid", fgColor=_COLOR_ENCABEZADO)
    fuente_encabezado = Font(bold=True, color="FFFFFF", size=8)
    fuente_datos = Font(size=8)
    alineacion = Alignment(horizontal="center", vertical="center")
    borde = _crear_borde_fino()

    # Encabezados de columnas de totales (fila 4)
    encabezados = ["Asignadas", "Objetivo", "Diferencia", "Estado"]
    for i, titulo in enumerate(encabezados):
        celda = ws.cell(row=4, column=col_inicio + i, value=titulo)
        celda.fill = estilo_encabezado
        celda.font = fuente_encabezado
        celda.alignment = alineacion
        celda.border = borde

    # Datos por asignatura (con fórmulas)
    fila_inicio = 5
    for idx, asignatura in enumerate(asignaturas):
        fila = fila_inicio + idx

        # Columna Asignadas: fórmula SUMA de las horas en la fila
        letra_asignadas = get_column_letter(col_asignadas)
        formula_suma = f"=SUM({col_primera_fecha}{fila}:{col_ultima_fecha}{fila})"
        celda_asignadas = ws.cell(row=fila, column=col_asignadas, value=formula_suma)
        celda_asignadas.font = fuente_datos
        celda_asignadas.alignment = alineacion
        celda_asignadas.border = borde

        # Columna Objetivo: valor fijo (las horas objetivo de la asignatura)
        horas_objetivo = asignatura.horas_totales
        celda_objetivo = ws.cell(row=fila, column=col_objetivo, value=horas_objetivo)
        celda_objetivo.font = fuente_datos
        celda_objetivo.alignment = alineacion
        celda_objetivo.border = borde

        # Columna Diferencia: fórmula (Asignadas - Objetivo)
        letra_objetivo = get_column_letter(col_objetivo)
        formula_diferencia = f"={letra_asignadas}{fila}-{letra_objetivo}{fila}"
        celda_diferencia = ws.cell(row=fila, column=col_diferencia, value=formula_diferencia)
        celda_diferencia.font = fuente_datos
        celda_diferencia.alignment = alineacion
        celda_diferencia.border = borde

        # Columna Estado: fórmula SI (Incompleta/Excede/Completa)
        letra_diferencia = get_column_letter(col_diferencia)
        formula_estado = f'=IF({letra_diferencia}{fila}<-0.1,"Incompleta",IF({letra_diferencia}{fila}>0.1,"Excede","Completa"))'
        celda_estado = ws.cell(row=fila, column=col_estado, value=formula_estado)
        celda_estado.font = fuente_datos
        celda_estado.alignment = alineacion
        celda_estado.border = borde

        # Aplicar formato condicional manual basado en el valor actual (para visualización inicial)
        horas_asignadas_actual = sum(
            matriz_horas.get((asignatura.codigo, fecha), 0.0)
            for fecha in fechas
        )
        diferencia_actual = horas_asignadas_actual - horas_objetivo

        if diferencia_actual < -0.1:
            color_estado = _COLOR_RESUMEN_INCOMPLETA
        elif diferencia_actual > 0.1:
            color_estado = _COLOR_RESUMEN_EXCEDE
        else:
            color_estado = _COLOR_RESUMEN_OK

        celda_estado.fill = PatternFill(fill_type="solid", fgColor=color_estado)


# ---------------------------------------------------------------------------
# Formato y ajustes finales
# ---------------------------------------------------------------------------

def _crear_borde_fino() -> Border:
    """Crea un borde fino gris para las celdas."""
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _ajustar_dimensiones(
    ws,
    fechas: list[datetime.date],
    asignaturas: list[Asignatura],
) -> None:
    """
    Ajusta anchos de columna y altos de fila.

    Args:
        ws: hoja de trabajo de openpyxl.
        fechas: lista de fechas (determina número de columnas).
        asignaturas: lista de asignaturas (determina número de filas).
    """
    # Columna A: Tipo
    ws.column_dimensions["A"].width = 18

    # Columna B: Asignatura
    ws.column_dimensions["B"].width = 40

    # Columnas de fechas: angostas (empiezan en C)
    for col_idx in range(len(fechas)):
        letra = get_column_letter(col_idx + 3)
        ws.column_dimensions[letra].width = 6

    # Columnas de totales
    col_totales = len(fechas) + 3
    for i in range(4):
        letra = get_column_letter(col_totales + i)
        ws.column_dimensions[letra].width = 10


def _congelar_paneles(ws) -> None:
    """
    Congela paneles para mantener visibles las filas auxiliares y las columnas A-B.

    Congela en C5: filas 1-4 (auxiliares + encabezado) y columnas A-B (Tipo, Asignatura).
    """
    ws.freeze_panes = "C5"

