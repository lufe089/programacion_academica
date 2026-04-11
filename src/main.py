"""
main.py — Punto de entrada del sistema de programación académica.

Orquesta la lectura del Excel, la construcción del calendario,
la asignación de sesiones y la exportación de resultados.

Modos de ejecución:
    1. Generar programación automática (desde restricciones.xlsx)
    2. Regenerar desde matriz ajustada (desde inputs/programacion_matriz.xlsx)
"""

import os
import sys

import pandas as pd

from calendar_builder import construir_calendario
from excel_reader import leer_excel, parsear_catalogo, parsear_franjas, parsear_parametros
from exports_franjas import exportar_version_franjas
from exports_hours import exportar_version_horas
from exports_matriz import exportar_matriz_horas
from exports_visual import exportar_version_visual
from scheduler import asignar_sesiones, construir_candidatos, construir_candidatos_desde_matriz, filtrar_asignaturas_del_semestre

RUTA_EXCEL = os.path.join("../inputs", "restricciones.xlsx")
RUTA_MATRIZ_AJUSTADA = os.path.join("../inputs", "programacion_matriz.xlsx")


def main() -> None:
    """
    Muestra el menú principal y ejecuta la opción seleccionada.
    """
    print("=== Sistema de programación académica ===")
    print()
    print("Seleccione una opción:")
    print("  1. Generar programación automática")
    print("  2. Regenerar desde matriz ajustada (inputs/programacion_matriz.xlsx)")
    print("  0. Salir")
    print()

    opcion = input("Opción: ").strip()

    if opcion == "1":
        generar_programacion_automatica()
    elif opcion == "2":
        regenerar_desde_matriz_ajustada()
    elif opcion == "0":
        print("Saliendo...")
        return
    else:
        print("Opción no válida.")
        return


def generar_programacion_automatica() -> None:
    """
    Ejecuta el flujo principal del sistema:
    1. Lee y parsea el archivo Excel (catálogo, parámetros, franjas).
    2. Construye el calendario de slots disponibles.
    3. Filtra las asignaturas del semestre activo.
    4. Construye los candidatos (asignatura × slot válido).
    5. Asigna sesiones semana por semana.
    6. Exporta todos los archivos de salida.
    """
    print()
    print("--- Generando programación automática ---")
    print()

    # --- Lectura del Excel ---
    print(f"Leyendo archivo: {RUTA_EXCEL}")
    df_catalogo, df_parametros, df_franjas = leer_excel(RUTA_EXCEL)
    print("Archivo leído correctamente.")
    print()

    # --- Parseo de cada hoja ---
    franjas = parsear_franjas(df_franjas)
    parametros = parsear_parametros(df_parametros)
    asignaturas = parsear_catalogo(df_catalogo, franjas)

    # --- Resumen de lo cargado ---
    _imprimir_resumen_parametros(parametros)
    _imprimir_resumen_franjas(franjas)
    _imprimir_resumen_asignaturas(asignaturas)

    # --- Construcción del calendario ---
    if parametros.fin_clases is None:
        print("ADVERTENCIA: El parámetro 'FIN_CLASES' no está definido en el Excel.")
        print("Agrega FIN_CLASES a la hoja 'parametros' para construir el calendario.")
        return

    calendario = construir_calendario(parametros, franjas)
    _imprimir_resumen_calendario(calendario)

    # --- Preparación de candidatos ---
    asignaturas_del_semestre = filtrar_asignaturas_del_semestre(
        asignaturas,
        parametros.semestre_programacion,
    )
    candidatos = construir_candidatos(asignaturas_del_semestre, calendario)
    _imprimir_resumen_candidatos(asignaturas_del_semestre, candidatos)

    # --- Asignación de sesiones ---
    sesiones = asignar_sesiones(asignaturas_del_semestre, candidatos, franjas, parametros)
    _imprimir_resumen_sesiones(asignaturas_del_semestre, sesiones)

    # --- Exportación de la versión de horas ---
    ruta_horas = os.path.join("..", "outputs", "programacion_horas.xlsx")
    exportar_version_horas(
        sesiones=sesiones,
        asignaturas=asignaturas_del_semestre,
        ruta_salida=ruta_horas,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión de horas exportada a '{ruta_horas}'.")

    # --- Exportación de la versión de franjas ---
    ruta_franjas = os.path.join("..", "outputs", "programacion_franjas.xlsx")
    exportar_version_franjas(
        sesiones=sesiones,
        asignaturas=asignaturas_del_semestre,
        ruta_salida=ruta_franjas,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión de franjas exportada a '{ruta_franjas}'.")

    # --- Exportación de la versión visual ---
    ruta_visual = os.path.join("..", "outputs", "programacion_visual.xlsx")
    exportar_version_visual(
        sesiones=sesiones,
        franjas=franjas,
        ruta_salida=ruta_visual,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión visual exportada a '{ruta_visual}'.")

    # --- Exportación de la matriz de horas ---
    ruta_matriz = os.path.join("..", "outputs", "programacion_matriz.xlsx")
    exportar_matriz_horas(
        sesiones=sesiones,
        asignaturas=asignaturas_del_semestre,
        parametros=parametros,
        calendario=calendario,
        ruta_salida=ruta_matriz,
    )
    print(f"  Matriz de horas exportada a '{ruta_matriz}'.")


def regenerar_desde_matriz_ajustada() -> None:
    """
    Regenera los archivos de salida a partir de una matriz ajustada manualmente.

    1. Lee el archivo de restricciones para obtener parámetros y franjas.
    2. Lee la matriz ajustada de inputs/programacion_matriz.xlsx.
    3. Convierte la matriz en sesiones.
    4. Regenera todos los archivos de salida.
    """
    print()
    print("--- Regenerando desde matriz ajustada ---")
    print()

    # Verificar que existe la matriz ajustada
    if not os.path.exists(RUTA_MATRIZ_AJUSTADA):
        print(f"ERROR: No se encontró el archivo '{RUTA_MATRIZ_AJUSTADA}'")
        print("Copie el archivo programacion_matriz.xlsx de outputs/ a inputs/ y ajústelo manualmente.")
        return

    # --- Lectura del Excel de restricciones (para parámetros y franjas) ---
    print(f"Leyendo archivo de restricciones: {RUTA_EXCEL}")
    df_catalogo, df_parametros, df_franjas = leer_excel(RUTA_EXCEL)

    franjas = parsear_franjas(df_franjas)
    parametros = parsear_parametros(df_parametros)
    asignaturas = parsear_catalogo(df_catalogo, franjas)

    asignaturas_del_semestre = filtrar_asignaturas_del_semestre(
        asignaturas,
        parametros.semestre_programacion,
    )

    # --- Leer la matriz ajustada ---
    # Primero extraemos las fechas con horas de la matriz
    print(f"Leyendo matriz ajustada: {RUTA_MATRIZ_AJUSTADA}")
    fechas_con_horas = _extraer_fechas_de_matriz(RUTA_MATRIZ_AJUSTADA, asignaturas_del_semestre, parametros)

    if not fechas_con_horas:
        print("ERROR: No se pudieron extraer fechas de la matriz.")
        return

    # Construir candidatos SOLO para las fechas presentes en la matriz
    candidatos = construir_candidatos_desde_matriz(asignaturas_del_semestre, fechas_con_horas, franjas)
    print(f"  Candidatos construidos: {len(candidatos)} (solo para fechas de la matriz)")

    # Ahora convertir la matriz en sesiones usando esos candidatos
    sesiones = _leer_matriz_ajustada(RUTA_MATRIZ_AJUSTADA, asignaturas_del_semestre, candidatos, franjas, parametros)

    if sesiones.empty:
        print("ERROR: No se pudieron leer sesiones de la matriz.")
        return

    print(f"  Se leyeron {len(sesiones)} sesiones de la matriz.")
    print()

    _imprimir_resumen_sesiones(asignaturas_del_semestre, sesiones)

    # --- Exportar archivos de salida ---
    ruta_horas = os.path.join("..", "outputs", "programacion_horas.xlsx")
    exportar_version_horas(
        sesiones=sesiones,
        asignaturas=asignaturas_del_semestre,
        ruta_salida=ruta_horas,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión de horas exportada a '{ruta_horas}'.")

    ruta_franjas_out = os.path.join("..", "outputs", "programacion_franjas.xlsx")
    exportar_version_franjas(
        sesiones=sesiones,
        asignaturas=asignaturas_del_semestre,
        ruta_salida=ruta_franjas_out,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión de franjas exportada a '{ruta_franjas_out}'.")

    ruta_visual = os.path.join("..", "outputs", "programacion_visual.xlsx")
    exportar_version_visual(
        sesiones=sesiones,
        franjas=franjas,
        ruta_salida=ruta_visual,
        inicio_clases=parametros.inicio_clases,
    )
    print(f"  Versión visual exportada a '{ruta_visual}'.")

    # La matriz NO se regenera (es la fuente de los ajustes)
    print()
    print("NOTA: La matriz de horas NO se regenera (es la fuente de los ajustes).")
    print("      Si desea regenerarla, ejecute la opción 1 nuevamente.")


def _extraer_fechas_de_matriz(
    ruta: str,
    asignaturas: list,
    parametros,
) -> dict[str, set]:
    """
    Extrae las fechas con horas > 0 de la matriz para cada asignatura.

    Lee la matriz y retorna un diccionario {codigo_asignatura: {fechas}}.
    Esto permite construir candidatos SOLO para las fechas que realmente
    tienen horas asignadas en la matriz.

    Args:
        ruta: ruta al archivo de matriz ajustada.
        asignaturas: lista de asignaturas del semestre.
        parametros: parámetros con inicio_clases para inferir el año.

    Returns:
        Diccionario {codigo: {fecha1, fecha2, ...}} con las fechas que tienen horas.
    """
    from datetime import datetime
    from openpyxl import load_workbook

    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    # Inferir año desde parámetros
    anio = parametros.inicio_clases.year

    # --- Leer fechas de la fila 3 (formato dd/mm) ---
    fechas_por_columna = {}
    for col in range(3, ws.max_column + 1):
        valor = ws.cell(row=3, column=col).value
        if valor and isinstance(valor, str) and "/" in valor:
            try:
                dia, mes = valor.split("/")
                fecha = datetime(anio, int(mes), int(dia)).date()
                fechas_por_columna[col] = fecha
            except (ValueError, AttributeError):
                break
        elif valor and hasattr(valor, "date"):
            fechas_por_columna[col] = valor.date()
        else:
            break

    if not fechas_por_columna:
        wb.close()
        return {}

    # --- Mapa nombre → asignatura ---
    asignatura_por_nombre = {a.nombre: a for a in asignaturas}

    # --- Leer filas de asignaturas (desde fila 5) ---
    fechas_con_horas: dict[str, set] = {}
    fila = 5

    while fila <= ws.max_row:
        nombre = ws.cell(row=fila, column=2).value

        if not nombre or nombre == "Horas efectivas":
            break

        asignatura = asignatura_por_nombre.get(nombre)
        if not asignatura:
            for a in asignaturas:
                if a.nombre in str(nombre) or str(nombre) in a.nombre:
                    asignatura = a
                    break

        if asignatura:
            codigo = asignatura.codigo
            if codigo not in fechas_con_horas:
                fechas_con_horas[codigo] = set()

            for col, fecha in fechas_por_columna.items():
                valor = ws.cell(row=fila, column=col).value
                if valor and isinstance(valor, (int, float)) and valor > 0:
                    fechas_con_horas[codigo].add(fecha)

        fila += 1

    wb.close()
    return fechas_con_horas


def _leer_matriz_ajustada(
    ruta: str,
    asignaturas: list,
    candidatos: pd.DataFrame,
    franjas: list,
    parametros,
) -> pd.DataFrame:
    """
    Lee la matriz ajustada y la convierte en un DataFrame de sesiones.

    Para cada (asignatura, fecha) con horas > 0 en la matriz, busca los slots
    candidatos válidos —que ya respetan las franjas permitidas y el calendario—
    y distribuye las horas en orden de franja.

    La matriz tiene:
    - Fila 1: números de semana (S1, S2, ...)
    - Fila 2: días de la semana (Mié, Vie, Sáb)
    - Fila 3: fechas (dd/mm)
    - Fila 4: encabezados (Tipo, Asignatura, fechas..., Asignadas, Objetivo, etc.)
    - Filas 5+: datos de asignaturas

    Args:
        ruta: ruta al archivo de matriz ajustada.
        asignaturas: lista de asignaturas del semestre.
        candidatos: DataFrame de slots válidos (asignatura × slot), construido
            desde las fechas de la matriz y las franjas permitidas.
        franjas: lista de franjas en orden de definición (determina prioridad).
        parametros: parámetros con inicio_clases para inferir el año.

    Returns:
        DataFrame con las sesiones leídas de la matriz.
    """
    from datetime import datetime
    from openpyxl import load_workbook

    # Inferir año desde parámetros
    anio = parametros.inicio_clases.year

    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    # --- Leer fechas de la fila 3 (formato dd/mm) ---
    fechas_por_columna = {}
    for col in range(3, ws.max_column + 1):
        valor = ws.cell(row=3, column=col).value
        if valor and isinstance(valor, str) and "/" in valor:
            try:
                dia, mes = valor.split("/")
                fecha = datetime(anio, int(mes), int(dia)).date()
                fechas_por_columna[col] = fecha
            except (ValueError, AttributeError):
                break
        elif valor and hasattr(valor, "date"):
            fechas_por_columna[col] = valor.date()
        else:
            break

    if not fechas_por_columna:
        print("ERROR: No se pudieron leer las fechas de la matriz.")
        return pd.DataFrame()

    # --- Índice de candidatos por (codigo, fecha) para búsqueda rápida ---
    # candidatos ya tiene hora_inicio, hora_fin, duracion_mins y respeta las
    # franjas permitidas de cada asignatura y los slots programables del calendario.
    franja_orden = {f.nombre: idx for idx, f in enumerate(franjas)}

    candidatos_por_clave: dict = {}
    if not candidatos.empty:
        for (codigo, fecha), grupo in candidatos.groupby(["codigo", "fecha"]):
            slots_ordenados = grupo.sort_values(
                "nombre_franja",
                key=lambda s: s.map(franja_orden).fillna(999),
            )
            candidatos_por_clave[(codigo, fecha)] = slots_ordenados

    # --- Mapa nombre → asignatura ---
    asignatura_por_nombre = {a.nombre: a for a in asignaturas}

    dias_map = {2: "miercoles", 4: "viernes", 5: "sabado"}

    # --- Registro de franjas ocupadas por tipo ---
    # Clave: (fecha, nombre_franja) -> tipo de asignatura que la ocupa
    # Esto evita mezclar tipos diferentes en la misma franja/fecha
    franjas_ocupadas: dict[tuple, str] = {}

    # --- Leer filas de asignaturas (desde fila 5) ---
    sesiones = []
    fila = 5

    while fila <= ws.max_row:
        nombre = ws.cell(row=fila, column=2).value
        tipo = ws.cell(row=fila, column=1).value  # Leer tipo de columna A

        if not nombre or nombre == "Horas efectivas":
            break

        asignatura = asignatura_por_nombre.get(nombre)
        if not asignatura:
            for a in asignaturas:
                if a.nombre in str(nombre) or str(nombre) in a.nombre:
                    asignatura = a
                    break

        # Si no se encontró la asignatura en el catálogo, es una fila adicional
        # Se procesa con información parcial (sin candidatos predefinidos)
        es_asignatura_adicional = asignatura is None

        if es_asignatura_adicional:
            print(f"  INFO: Asignatura adicional detectada: '{nombre}' (tipo: {tipo})")

        for col, fecha in fechas_por_columna.items():
            valor = ws.cell(row=fila, column=col).value
            if not valor or not isinstance(valor, (int, float)) or valor <= 0:
                continue

            horas = float(valor)
            dia_nombre = dias_map.get(fecha.weekday(), "")

            if es_asignatura_adicional:
                # Para asignaturas adicionales: crear sesión con campos parciales
                # Los campos que no conocemos se dejan vacíos o con valores por defecto
                sesiones.append({
                    "codigo": "",  # Sin código conocido
                    "asignatura": nombre,
                    "tipo": tipo if tipo else "",
                    "fecha": fecha,
                    "dia_semana": dia_nombre,
                    "nombre_franja": "",  # Sin franja conocida
                    "hora_inicio": None,
                    "hora_fin": None,
                    "duracion_mins": None,
                    "horas_sesion": horas,
                    "horas_acumuladas": 0.0,
                })
            else:
                # Para asignaturas del catálogo: usar candidatos respetando tipos
                slots = candidatos_por_clave.get((asignatura.codigo, fecha))
                if slots is None or slots.empty:
                    print(
                        f"  ADVERTENCIA: Sin candidatos para {asignatura.codigo} "
                        f"en {fecha} — se omite la fecha."
                    )
                    continue

                horas_restantes = horas
                tipo_asignatura = asignatura.tipo

                for _, slot in slots.iterrows():
                    if horas_restantes <= 0:
                        break

                    nombre_franja = slot["nombre_franja"]
                    clave_franja = (fecha, nombre_franja)

                    # Verificar si la franja ya está ocupada por otro tipo
                    tipo_ocupante = franjas_ocupadas.get(clave_franja)
                    if tipo_ocupante is not None and tipo_ocupante != tipo_asignatura:
                        # La franja tiene otro tipo, saltar a la siguiente
                        continue

                    horas_franja = min(slot["duracion_mins"] / 60, horas_restantes)

                    sesiones.append({
                        "codigo": asignatura.codigo,
                        "asignatura": asignatura.nombre,
                        "tipo": tipo_asignatura,
                        "fecha": fecha,
                        "dia_semana": dia_nombre,
                        "nombre_franja": nombre_franja,
                        "hora_inicio": slot["hora_inicio"],
                        "hora_fin": slot["hora_fin"],
                        "duracion_mins": slot["duracion_mins"],
                        "horas_sesion": horas_franja,
                        "horas_acumuladas": 0.0,
                    })

                    # Registrar la franja como ocupada por este tipo
                    franjas_ocupadas[clave_franja] = tipo_asignatura

                    horas_restantes -= horas_franja

                if horas_restantes > 0:
                    print(
                        f"  ADVERTENCIA: {asignatura.codigo} en {fecha} tiene "
                        f"{horas_restantes:.1f}h sin asignar (franjas ocupadas por otros tipos)."
                    )

        fila += 1

    wb.close()

    if not sesiones:
        return pd.DataFrame()

    df = pd.DataFrame(sesiones)
    # Ordenar por asignatura (nombre si código vacío) y fecha
    df = df.sort_values(by=["asignatura", "fecha"]).reset_index(drop=True)

    # Calcular horas acumuladas por asignatura (usando nombre como clave)
    for nombre_asig in df["asignatura"].unique():
        mask = df["asignatura"] == nombre_asig
        df.loc[mask, "horas_acumuladas"] = df.loc[mask, "horas_sesion"].cumsum()

    return df


def _imprimir_resumen_parametros(parametros) -> None:
    """Imprime los parámetros cargados de la corrida."""
    print("--- Parámetros de la corrida ---")
    print(f"  Semestre: {parametros.semestre_programacion}")
    print(f"  Fecha inducción: {parametros.fecha_induccion}")
    print(f"  Inicio clases: {parametros.inicio_clases}")
    print(f"  Fin clases: {parametros.fin_clases if parametros.fin_clases else 'No definido'}")
    print(f"  Semana sin clases: {parametros.inicio_semana_sin_clases} al {parametros.fin_semana_sin_clases}")
    print(f"  Festivos: {parametros.festivos}")
    print()


def _imprimir_resumen_franjas(franjas) -> None:
    """Imprime las franjas cargadas."""
    print(f"--- Franjas definidas ({len(franjas)}) ---")
    for franja in franjas:
        print(
            f"  {franja.nombre:<25} "
            f"{franja.dia_semana.value:<10} "
            f"{franja.hora_inicio} – {franja.hora_fin} "
            f"({franja.duracion_minutos} min)"
        )
    print()


def _imprimir_resumen_asignaturas(asignaturas) -> None:
    """Imprime las asignaturas cargadas del catálogo."""
    print(f"--- Asignaturas en el catálogo ({len(asignaturas)}) ---")
    for asignatura in asignaturas:
        print(
            f"  [{asignatura.codigo}] {asignatura.nombre[:55]:<55} "
            f"| {asignatura.semestre_oferta.value:<8} "
            f"| {asignatura.horas_totales}h "
            f"| {asignatura.restriccion_programacion.value}"
        )
    print()


def _imprimir_resumen_calendario(calendario) -> None:
    """Imprime estadísticas del calendario construido."""
    total_slots = len(calendario)
    slots_disponibles = calendario["es_programable"].sum()
    slots_bloqueados = total_slots - slots_disponibles

    print("--- Calendario del periodo ---")
    print(f"  Total de slots (fecha × franja): {total_slots}")
    print(f"  Slots disponibles para programar: {slots_disponibles}")
    print(f"  Slots bloqueados: {slots_bloqueados}")

    if slots_bloqueados > 0:
        print()
        print("  Detalle de bloqueos:")
        bloqueados = calendario[~calendario["es_programable"]]
        por_motivo = bloqueados.groupby("motivo_bloqueo")["fecha"].count()
        for motivo, cantidad in por_motivo.items():
            print(f"    {motivo}: {cantidad} slots")

    print()
    print("  Primeras filas del calendario:")
    print(calendario.head(10).to_string(index=False))
    calendario.to_excel("calendario.xlsx", index=False)
    print("  Calendario exportado a 'calendario.xlsx'.")


def _imprimir_resumen_sesiones(asignaturas, sesiones) -> None:
    """Imprime el resumen de horas asignadas vs esperadas por asignatura."""
    print("--- Resumen de sesiones asignadas ---")
    for asignatura in asignaturas:
        sesiones_asignatura = sesiones[sesiones["codigo"] == asignatura.codigo]
        horas_asignadas = sesiones_asignatura["horas_sesion"].sum()
        semanas_usadas = sesiones_asignatura["fecha"].nunique()
        faltante = asignatura.horas_totales - horas_asignadas
        estado = "OK" if faltante == 0 else f"FALTA {faltante:.1f}h"
        print(
            f"  [{asignatura.codigo}] {asignatura.nombre[:45]:<45} "
            f"| {horas_asignadas:.1f}/{asignatura.horas_totales}h "
            f"| {semanas_usadas} semanas "
            f"| {estado}"
        )
    print()


def _imprimir_resumen_candidatos(asignaturas_del_semestre, candidatos) -> None:
    """Imprime estadísticas de las asignaturas y candidatos de la corrida."""
    print(f"--- Asignaturas de la corrida ({len(asignaturas_del_semestre)}) ---")
    for asignatura in asignaturas_del_semestre:
        slots_asignatura = candidatos[candidatos["codigo"] == asignatura.codigo]
        print(
            f"  [{asignatura.codigo}] {asignatura.nombre[:50]:<50} "
            f"| {asignatura.horas_totales}h "
            f"| {len(slots_asignatura)} slots candidatos"
        )
    print()
    print(f"  Total de candidatos (asignatura × slot): {len(candidatos)}")
    print()


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, ValueError) as error:
        print(f"\nError: {error}", file=sys.stderr)
        sys.exit(1)
