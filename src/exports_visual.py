"""
exports_visual.py — Exportación de la versión gráfica de la programación.

Genera un archivo Excel con una cuadrícula de calendario donde:

    - Filas: franjas horarias definidas (en orden cronológico).
    - Columnas: semanas del semestre (con número y fecha de inicio).
    - Celdas: código(s) de asignatura programados en esa franja y semana.

Cada tipo de asignatura recibe un color de fondo distinto para facilitar
la lectura visual de la carga semanal y la distribución por tipo.

Este módulo solo se ocupa de la presentación y formato del output.
No contiene lógica de negocio ni de asignación.
"""

import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exports_hours import calcular_numero_semana
from models import Franja


# ---------------------------------------------------------------------------
# Colores de fondo por tipo de asignatura (hex RGB sin #)
# ---------------------------------------------------------------------------

_COLORES_POR_TIPO: dict[str, str] = {
    "Obligatorio": "FFA420",        # Naranja claro
    "TemasAvanzados": "CCA9DD",     # purpura claro
    "ProcesoDesarrollo": "A8D8EA",  # azul claro
    "ProyectoGrado": "B8E0B0",      # verde claro
    "SoloMiercoles": "B8E0B0",      # verde claro (mismo que ProyectoGrado)
}
_COLOR_DEFAULT = "E8E8E8"           # gris claro (tipo no reconocido)
_COLOR_VACIO = "FFFFFF"             # blanco (sin asignación)
_COLOR_ENCABEZADO = "2F4F8F"        # azul oscuro para encabezados
_COLOR_FRANJA_LABEL = "F2F2F2"      # gris muy claro para etiquetas de franja


# ---------------------------------------------------------------------------
# Función principal de exportación
# ---------------------------------------------------------------------------

def exportar_version_visual(
    sesiones: pd.DataFrame,
    franjas: list[Franja],
    ruta_salida: str,
    inicio_clases: datetime.date,
) -> None:
    """
    Exporta la programación en formato de cuadrícula de calendario.

    Genera un archivo Excel con una hoja 'calendario' que muestra,
    para cada semana del semestre, qué asignaturas están programadas
    en cada franja horaria.

    Args:
        sesiones: DataFrame de sesiones asignadas (salida de asignar_sesiones).
        franjas: lista de objetos Franja en orden de definición del Excel.
        ruta_salida: ruta completa del archivo Excel a generar.
        inicio_clases: fecha de inicio de clases, usada para numerar semanas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "calendario"

    if sesiones.empty:
        wb.save(ruta_salida)
        return

    df = sesiones.copy()
    df["semana"] = df["fecha"].apply(
        lambda fecha: calcular_numero_semana(fecha, inicio_clases)
    )

    semanas_ordenadas = sorted(df["semana"].unique())
    lunes_por_semana = _calcular_lunes_por_semana(df, semanas_ordenadas)

    _escribir_encabezados(ws, franjas, semanas_ordenadas, lunes_por_semana)
    _escribir_celdas(ws, df, franjas, semanas_ordenadas)
    _ajustar_dimensiones(ws, semanas_ordenadas)

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# Construcción de la cuadrícula
# ---------------------------------------------------------------------------

def _calcular_lunes_por_semana(
    df: pd.DataFrame,
    semanas: list[int],
) -> dict[int, datetime.date]:
    """
    Calcula la fecha del lunes para cada número de semana.

    Args:
        df: DataFrame de sesiones con columna 'semana' ya calculada.
        semanas: lista de números de semana presentes en las sesiones.

    Returns:
        Diccionario semana → fecha del lunes de esa semana.
    """
    lunes_por_semana: dict[int, datetime.date] = {}
    for semana in semanas:
        fecha_muestra = df[df["semana"] == semana]["fecha"].iloc[0]
        lunes = fecha_muestra - datetime.timedelta(days=fecha_muestra.weekday())
        lunes_por_semana[semana] = lunes
    return lunes_por_semana


def _escribir_encabezados(
    ws,
    franjas: list[Franja],
    semanas: list[int],
    lunes_por_semana: dict[int, datetime.date],
) -> None:
    """
    Escribe la fila de encabezado con los números de semana y fechas de inicio.

    Columna 1: etiqueta "Franja / Semana"
    Columnas 2+: una por semana con "S{n}\n{dd-MMM}" como texto.

    Args:
        ws: hoja de trabajo de openpyxl.
        franjas: lista de franjas (define el número de filas de datos).
        semanas: lista de números de semana en orden.
        lunes_por_semana: fecha del lunes por cada semana.
    """
    estilo_encabezado = PatternFill(fill_type="solid", fgColor=_COLOR_ENCABEZADO)
    fuente_encabezado = Font(bold=True, color="FFFFFF", size=9)
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Celda de esquina
    celda_esquina = ws.cell(row=1, column=1, value="Franja / Semana")
    celda_esquina.fill = estilo_encabezado
    celda_esquina.font = fuente_encabezado
    celda_esquina.alignment = alineacion_centro

    for idx, semana in enumerate(semanas):
        columna = idx + 2
        lunes = lunes_por_semana[semana]
        texto = f"S{semana}\n{lunes.strftime('%d-%b')}"
        celda = ws.cell(row=1, column=columna, value=texto)
        celda.fill = estilo_encabezado
        celda.font = fuente_encabezado
        celda.alignment = alineacion_centro


def _escribir_celdas(
    ws,
    df: pd.DataFrame,
    franjas: list[Franja],
    semanas: list[int],
) -> None:
    """
    Escribe las filas de datos con etiquetas de franja y asignaturas por semana.

    Para cada (franja, semana), muestra los códigos de asignatura programados.
    El color de fondo se determina por el tipo de la(s) asignatura(s).

    Args:
        ws: hoja de trabajo de openpyxl.
        df: DataFrame de sesiones con columna 'semana'.
        franjas: lista de objetos Franja en orden cronológico.
        semanas: lista de números de semana en orden.
    """
    estilo_franja_label = PatternFill(fill_type="solid", fgColor=_COLOR_FRANJA_LABEL)
    fuente_label = Font(bold=True, size=8)
    fuente_datos = Font(size=8)
    alineacion_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alineacion_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde_fino = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for fila_idx, franja in enumerate(franjas):
        fila = fila_idx + 2

        # Celda de etiqueta de franja
        etiqueta = _formatear_etiqueta_franja(franja)
        celda_label = ws.cell(row=fila, column=1, value=etiqueta)
        celda_label.fill = estilo_franja_label
        celda_label.font = fuente_label
        celda_label.alignment = alineacion_izq
        celda_label.border = borde_fino

        # Celdas de datos por semana
        for col_idx, semana in enumerate(semanas):
            columna = col_idx + 2
            sesiones_slot = df[
                (df["nombre_franja"] == franja.nombre) & (df["semana"] == semana)
            ]

            if sesiones_slot.empty:
                color_fondo = _COLOR_VACIO
                texto_celda = ""
            else:
                color_fondo = _determinar_color_slot(sesiones_slot)
                texto_celda = _formatear_contenido_celda(sesiones_slot)

            celda = ws.cell(row=fila, column=columna, value=texto_celda)
            celda.fill = PatternFill(fill_type="solid", fgColor=color_fondo)
            celda.font = fuente_datos
            celda.alignment = alineacion_centro
            celda.border = borde_fino


def _formatear_etiqueta_franja(franja: Franja) -> str:
    """
    Construye la etiqueta de texto para la columna de nombre de franja.

    Formato: "NOMBRE_FRANJA\nHH:MM – HH:MM"

    Args:
        franja: objeto Franja con nombre y horarios.

    Returns:
        Texto formateado para mostrar en la celda de etiqueta.
    """
    hora_ini = franja.hora_inicio.strftime("%H:%M")
    hora_fin = franja.hora_fin.strftime("%H:%M")
    return f"{franja.nombre}\n{hora_ini} – {hora_fin}"


def _determinar_color_slot(sesiones_slot: pd.DataFrame) -> str:
    """
    Determina el color de fondo de una celda según el tipo de las asignaturas.

    Si todas las asignaturas del slot son del mismo tipo, usa el color de ese tipo.
    Si hay tipos mezclados, usa el color por defecto.

    Args:
        sesiones_slot: sesiones programadas en esa (franja, semana).

    Returns:
        Código de color hexadecimal (sin #).
    """
    tipos_presentes = sesiones_slot["tipo"].unique()
    if len(tipos_presentes) == 1:
        tipo = tipos_presentes[0]
        return _COLORES_POR_TIPO.get(tipo, _COLOR_DEFAULT)
    return _COLOR_DEFAULT


def _formatear_contenido_celda(sesiones_slot: pd.DataFrame) -> str:
    """
    Construye el texto de una celda con los códigos de asignatura presentes.

    Para cada asignatura única en el slot, muestra los últimos 6 caracteres
    del código. Si hay varias, las separa con saltos de línea.

    Args:
        sesiones_slot: sesiones programadas en esa (franja, semana).

    Returns:
        Texto con los códigos de asignatura, separados por salto de línea.
    """
    codigos = sesiones_slot["codigo"].unique()
    etiquetas = [codigo[-6:] for codigo in sorted(codigos)]
    return "\n".join(etiquetas)


# ---------------------------------------------------------------------------
# Ajuste de dimensiones de la hoja
# ---------------------------------------------------------------------------

def _ajustar_dimensiones(ws, semanas: list[int]) -> None:
    """
    Ajusta el ancho de columnas y el alto de filas para una lectura cómoda.

    Args:
        ws: hoja de trabajo de openpyxl.
        semanas: lista de semanas, usada para saber cuántas columnas de datos hay.
    """
    ws.column_dimensions["A"].width = 28
    ancho_semana = 11
    for idx in range(len(semanas)):
        letra = get_column_letter(idx + 2)
        ws.column_dimensions[letra].width = ancho_semana

    ws.row_dimensions[1].height = 30
    for fila in range(2, ws.max_row + 1):
        ws.row_dimensions[fila].height = 30

    ws.freeze_panes = "B2"
